"""
Serialize analytics dashboard payloads for CSV, Excel, and PDF download.
"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook


def render_dashboard_export_csv(dashboard_payload: dict) -> str:
    """Same layout as the legacy single-sheet CSV export."""
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["Metric", "Value"])
    for key, value in dashboard_payload["metrics"].items():
        writer.writerow([key, value])

    writer.writerow([])
    writer.writerow(["Application Status", "Count"])
    for key, value in dashboard_payload["application_status"].items():
        writer.writerow([key, value])

    writer.writerow([])
    writer.writerow(
        ["Program", "Applications", "Approval Rate", "Avg Processing Time", "Popularity Score"]
    )
    for program in dashboard_payload["program_performance"]:
        writer.writerow(
            [
                program["name"],
                program["applications"],
                program["approval_rate"],
                program["avg_processing_time"],
                program["popularity_score"],
            ]
        )

    return output.getvalue()


def render_dashboard_export_xlsx(dashboard_payload: dict) -> bytes:
    """Multi-sheet workbook mirroring the CSV sections."""
    wb = Workbook()

    ws_m = wb.active
    ws_m.title = "Metrics"
    ws_m.append(["Metric", "Value"])
    for key, value in dashboard_payload["metrics"].items():
        ws_m.append([key, value])

    ws_s = wb.create_sheet("Application status")
    ws_s.append(["Application Status", "Count"])
    for key, value in dashboard_payload["application_status"].items():
        ws_s.append([key, value])

    ws_p = wb.create_sheet("Program performance")
    ws_p.append(
        ["Program", "Applications", "Approval Rate", "Avg Processing Time", "Popularity Score"]
    )
    for program in dashboard_payload["program_performance"]:
        ws_p.append(
            [
                program["name"],
                program["applications"],
                program["approval_rate"],
                program["avg_processing_time"],
                program["popularity_score"],
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def render_dashboard_export_pdf(dashboard_payload: dict) -> bytes:
    """Single PDF (landscape) with the same three sections as CSV/Excel."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    page = landscape(letter)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        rightMargin=36,
        leftMargin=36,
        topMargin=42,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("SEIM Analytics Report", styles["Title"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Metrics", styles["Heading2"]))
    m_data = [["Metric", "Value"]] + [
        [str(k), str(v)] for k, v in dashboard_payload["metrics"].items()
    ]
    tw = page[0] - 72
    t1 = Table(m_data, colWidths=[tw * 0.55, tw * 0.45])
    t1.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]
        )
    )
    story.append(t1)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Application status", styles["Heading2"]))
    s_data = [["Application status", "Count"]] + [
        [str(k), str(v)] for k, v in dashboard_payload["application_status"].items()
    ]
    t2 = Table(s_data, colWidths=[tw * 0.55, tw * 0.45])
    t2.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]
        )
    )
    story.append(t2)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Program performance", styles["Heading2"]))
    p_header = [
        "Program",
        "Applications",
        "Approval %",
        "Avg proc. (d)",
        "Popularity",
    ]
    p_rows = [p_header]
    for p in dashboard_payload["program_performance"]:
        name = str(p.get("name", ""))
        if len(name) > 72:
            name = name[:69] + "..."
        p_rows.append(
            [
                name,
                str(p.get("applications", "")),
                str(p.get("approval_rate", "")),
                str(p.get("avg_processing_time", "")),
                str(p.get("popularity_score", "")),
            ]
        )
    col_w = tw / 5
    t3 = Table(p_rows, colWidths=[col_w * 1.4, col_w * 0.9, col_w * 0.9, col_w * 0.9, col_w * 0.9])
    t3.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t3)

    doc.build(story)
    return buffer.getvalue()
