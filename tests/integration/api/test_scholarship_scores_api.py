"""API tests for scholarship allocation scoring (staff detail + cohort exports)."""

from django.urls import reverse
from rest_framework import status

from exchange.scholarship_scoring import STUDENT_SCHOLARSHIP_DISCLAIMER
from tests.utils import APITestCase


class TestScholarshipScoresAPI(APITestCase):
    def test_student_detail_includes_estimate_with_student_disclaimer(self):
        student = self.create_user(role="student")
        app = self.create_application(student=student, status_name="submitted")
        self.authenticate_user(student)
        url = reverse("api:application-detail", args=[app.id])
        response = self.client.get(url)
        self.assert_response_success(response, status.HTTP_200_OK)
        sc = response.data.get("scholarship_allocation_score")
        self.assertIsNotNone(sc)
        self.assertEqual(sc["ruleset_id"], "default_v1")
        self.assertEqual(len(sc["factors"]), 5)
        self.assertEqual(sc["disclaimer"], STUDENT_SCHOLARSHIP_DISCLAIMER)

    def test_coordinator_detail_includes_score(self):
        student = self.create_user(role="student")
        coordinator = self.create_user(role="coordinator")
        app = self.create_application(student=student, status_name="submitted")
        self.authenticate_user(coordinator)
        url = reverse("api:application-detail", args=[app.id])
        response = self.client.get(url)
        self.assert_response_success(response, status.HTTP_200_OK)
        sc = response.data.get("scholarship_allocation_score")
        self.assertIsNotNone(sc)
        self.assertEqual(sc["ruleset_id"], "default_v1")
        self.assertEqual(len(sc["factors"]), 5)

    def test_export_requires_program_param(self):
        coordinator = self.create_user(role="coordinator")
        self.authenticate_user(coordinator)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_export_student_forbidden(self):
        student = self.create_user(role="student")
        program = self.create_program()
        self.authenticate_user(student)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(url, {"program": str(program.id)})
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_export_coordinator_csv(self):
        program = self.create_program()
        student = self.create_user(role="student")
        self.create_application(student=student, program=program, status_name="submitted")
        coordinator = self.create_user(role="coordinator")
        self.authenticate_user(coordinator)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(url, {"program": str(program.id)})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        assert "rank" in body.splitlines()[0]
        assert "default_v1" in body

    def test_export_coordinator_xlsx(self):
        from io import BytesIO

        from openpyxl import load_workbook

        program = self.create_program()
        student = self.create_user(role="student")
        self.create_application(student=student, program=program, status_name="submitted")
        coordinator = self.create_user(role="coordinator")
        self.authenticate_user(coordinator)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(
            url, {"program": str(program.id), "export_format": "xlsx"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn(".xlsx", response["Content-Disposition"])
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(wb.sheetnames[0], "Scores")
        ws = wb.active
        self.assertEqual(ws["A1"].value, "rank")

    def test_export_coordinator_pdf(self):
        program = self.create_program()
        student = self.create_user(role="student")
        self.create_application(student=student, program=program, status_name="submitted")
        coordinator = self.create_user(role="coordinator")
        self.authenticate_user(coordinator)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(
            url, {"program": str(program.id), "export_format": "pdf"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("application/pdf", response["Content-Type"])
        self.assertIn(".pdf", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_export_invalid_format_returns_400(self):
        program = self.create_program()
        coordinator = self.create_user(role="coordinator")
        self.authenticate_user(coordinator)
        url = reverse("api:application-scholarship-scores-export")
        response = self.client.get(
            url, {"program": str(program.id), "export_format": "docx"}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
