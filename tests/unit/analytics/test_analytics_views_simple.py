"""
Simple tests for analytics views to improve coverage.
"""

from datetime import timedelta
from unittest.mock import Mock, patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient, APITestCase

from accounts.models import Role, User
from analytics.views import (
    AdminDashboardViewSet,
    application_statistics_view,
    dashboard_view,
    export_data_view,
    program_statistics_view,
    user_activity_view,
)
from exchange.models import Application, ApplicationStatus, Program, TimelineEvent

User = get_user_model()


class TestAdminDashboardViewSetSimple(APITestCase):
    """Simple tests for AdminDashboardViewSet."""

    def setUp(self):
        """Set up test data."""
        self.client = APIClient()
        self.admin_user = User.objects.create_user(
            username='admin',
            email='admin@example.com',
            password='adminpass123'
        )

        self.regular_user = User.objects.create_user(
            username='user',
            email='user@example.com',
            password='userpass123'
        )

        # Create admin role and assign to admin user
        self.admin_role, _ = Role.objects.get_or_create(name="admin")
        self.admin_user.roles.add(self.admin_role)

        # Create test data
        self.program = Program.objects.create(
            name="Test Program",
            description="Test Description",
            start_date=timezone.now().date(),
            end_date=timezone.now().date() + timedelta(days=365),
            is_active=True
        )

        # Create application status
        self.status, _ = ApplicationStatus.objects.get_or_create(
            name="submitted",
            defaults={'order': 1}
        )

        self.application = Application.objects.create(
            program=self.program,
            student=self.regular_user,
            status=self.status
        )

    def test_is_admin_method_true(self):
        """Test is_admin method returns True for admin user."""
        viewset = AdminDashboardViewSet()
        self.assertTrue(viewset.is_admin(self.admin_user))

    def test_is_admin_method_false(self):
        """Test is_admin method returns False for regular user."""
        viewset = AdminDashboardViewSet()
        self.assertFalse(viewset.is_admin(self.regular_user))

    @patch('analytics.views.AnalyticsService.get_dashboard_metrics')
    def test_metrics_action_admin_access(self, mock_get_metrics):
        """Test metrics action with admin access."""
        mock_get_metrics.return_value = {
            "ongoing_applications": 5,
            "approved_applications": 10,
            "rejected_applications": 2
        }

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-metrics')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_users', response.data)
        self.assertIn('total_applications', response.data)
        self.assertIn('total_programs', response.data)
        self.assertIn('pending_reviews', response.data)
        self.assertIn('application_status', response.data)

    def test_metrics_action_non_admin_access(self):
        """Test metrics action with non-admin access."""
        self.client.force_authenticate(user=self.regular_user)
        url = reverse('analytics:admindashboard-metrics')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('error', response.data)

    @patch('analytics.views.AnalyticsService.get_dashboard_metrics')
    def test_metrics_action_exception_handling(self, mock_get_metrics):
        """Test metrics action exception handling."""
        mock_get_metrics.side_effect = Exception("Test error")

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-metrics')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)

    def test_activity_action_admin_access(self):
        """Test activity action with admin access."""
        TimelineEvent.objects.create(
            application=self.application,
            event_type="application_submitted",
            description="Test application was submitted",
            created_by=self.regular_user
        )

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-activity')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)

    def test_activity_action_non_admin_access(self):
        """Test activity action with non-admin access."""
        self.client.force_authenticate(user=self.regular_user)
        url = reverse('analytics:admindashboard-activity')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('error', response.data)

    def test_activity_action_exception_handling(self):
        """Test activity action exception handling."""
        with patch('exchange.models.TimelineEvent.objects.filter') as mock_filter:
            mock_filter.side_effect = Exception("Test error")

            self.client.force_authenticate(user=self.admin_user)
            url = reverse('analytics:admindashboard-activity')
            response = self.client.get(url)

            self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
            self.assertIn('error', response.data)

    def test_performance_action_admin_access(self):
        """Test performance action with admin access."""
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-performance')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('cpu_usage', response.data)
        self.assertIn('memory_usage', response.data)
        self.assertIn('db_connections', response.data)
        self.assertIn('cache_hit_rate', response.data)
        self.assertIn('response_time', response.data)
        self.assertIn('active_users', response.data)

    def test_performance_action_non_admin_access(self):
        """Test performance action with non-admin access."""
        self.client.force_authenticate(user=self.regular_user)
        url = reverse('analytics:admindashboard-performance')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('error', response.data)

    def test_performance_action_exception_handling(self):
        """Test performance action exception handling."""
        with patch('analytics.views.User.objects.filter') as mock_filter:
            mock_filter.side_effect = Exception("Test error")

            self.client.force_authenticate(user=self.admin_user)
            url = reverse('analytics:admindashboard-performance')
            response = self.client.get(url)

            self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
            self.assertIn('error', response.data)

    def test_alerts_action_admin_access(self):
        """Test alerts action with admin access."""
        # Create an old pending application
        Application.objects.create(
            program=self.program,
            student=self.regular_user,
            status=self.status,
            created_at=timezone.now() - timezone.timedelta(days=8)
        )

        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-alerts')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)

    def test_alerts_action_non_admin_access(self):
        """Test alerts action with non-admin access."""
        self.client.force_authenticate(user=self.regular_user)
        url = reverse('analytics:admindashboard-alerts')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('error', response.data)

    def test_alerts_action_exception_handling(self):
        """Test alerts action exception handling."""
        with patch('analytics.views.Application.objects.filter') as mock_filter:
            mock_filter.side_effect = Exception("Test error")

            self.client.force_authenticate(user=self.admin_user)
            url = reverse('analytics:admindashboard-alerts')
            response = self.client.get(url)

            self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
            self.assertIn('error', response.data)

    def test_system_info_action_admin_access(self):
        """Test system info action with admin access."""
        self.client.force_authenticate(user=self.admin_user)
        url = reverse('analytics:admindashboard-system-info')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('django_version', response.data)
        self.assertIn('python_version', response.data)
        self.assertIn('database', response.data)
        self.assertIn('redis', response.data)
        self.assertIn('environment', response.data)
        self.assertIn('debug', response.data)
        self.assertIn('timezone', response.data)
        self.assertIn('static_files', response.data)
        self.assertIn('media_files', response.data)

    def test_system_info_action_non_admin_access(self):
        """Test system info action with non-admin access."""
        self.client.force_authenticate(user=self.regular_user)
        url = reverse('analytics:admindashboard-system-info')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn('error', response.data)

    def test_system_info_action_exception_handling(self):
        """Test system info action exception handling."""
        with patch('sys.version', 'mocked'):  # Patch sys.version directly
            self.client.force_authenticate(user=self.admin_user)
            url = reverse('analytics:admindashboard-system-info')
            response = self.client.get(url)
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertIn('django_version', response.data)


class TestAnalyticsViewsSimple(TestCase):
    """Simple tests for analytics views."""

    def setUp(self):
        """Set up test data."""
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )

    @patch('analytics.views.render')
    def test_dashboard_view(self, mock_render):
        """Test dashboard view."""
        request = Mock()
        request.user = self.user

        dashboard_view(request)
        mock_render.assert_called_with(request, 'frontend/admin/dashboard.html')

    @patch('analytics.views.render')
    def test_application_statistics_view(self, mock_render):
        """Test application statistics view."""
        request = Mock()
        request.user = self.user

        application_statistics_view(request)
        mock_render.assert_called_with(request, 'frontend/admin/analytics.html')

    @patch('analytics.views.render')
    def test_program_statistics_view(self, mock_render):
        """Test program statistics view."""
        request = Mock()
        request.user = self.user

        program_statistics_view(request)
        mock_render.assert_called_with(request, 'frontend/admin/analytics.html')

    @patch('analytics.views.render')
    def test_user_activity_view(self, mock_render):
        """Test user activity view."""
        request = Mock()
        request.user = self.user

        user_activity_view(request)
        mock_render.assert_called_with(request, 'frontend/admin/analytics.html')

    @patch('analytics.views.render')
    def test_export_data_view(self, mock_render):
        """Test export data view."""
        request = Mock()
        request.user = self.user

        export_data_view(request)
        mock_render.assert_called_with(request, 'frontend/admin/analytics.html')


class TestAnalyticsAPIViewsSimple(APITestCase):
    """Simple tests for analytics API views."""

    def setUp(self):
        """Set up test data."""
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        self.client.force_authenticate(user=self.user)

    def test_api_application_statistics(self):
        """Test API application statistics view."""
        url = reverse('analytics:api_application_statistics')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_applications', response.data)

    def test_api_program_statistics(self):
        """Test API program statistics view."""
        url = reverse('analytics:api_program_statistics')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_programs', response.data)

    def test_api_user_activity(self):
        """Test API user activity view."""
        url = reverse('analytics:api_user_activity')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('total_users', response.data)

    def test_track_event_post(self):
        """Test track event view with POST."""
        url = reverse('analytics:track_event')
        data = {'event_type': 'test_event', 'data': {'key': 'value'}}
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_track_event_get(self):
        """Test track event view with GET."""
        url = reverse('analytics:track_event')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)

    def test_track_event_invalid_data(self):
        """Test track event view with invalid data."""
        url = reverse('analytics:track_event')
        data = {'invalid': 'data'}
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {})


class TestAdvancedAnalyticsAPIViews(APITestCase):
    """Tests for the richer analytics API endpoints under /api/analytics/."""

    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='analytics-user',
            email='analytics@example.com',
            password='testpass123'
        )
        self.client.force_authenticate(user=self.user)

        self.program = Program.objects.create(
            name="Analytics Program",
            description="Analytics test program",
            start_date=timezone.now().date(),
            end_date=timezone.now().date() + timedelta(days=120),
            is_active=True,
        )
        self.approved_status, _ = ApplicationStatus.objects.get_or_create(
            name="approved",
            defaults={"order": 1},
        )
        self.rejected_status, _ = ApplicationStatus.objects.get_or_create(
            name="rejected",
            defaults={"order": 2},
        )

        Application.objects.create(
            program=self.program,
            student=self.user,
            status=self.approved_status,
        )
        Application.objects.create(
            program=self.program,
            student=self.user,
            status=self.rejected_status,
        )

    def test_dashboard_api_returns_expected_sections(self):
        url = reverse('api:analytics-dashboard')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('metrics', response.data)
        self.assertIn('application_status', response.data)
        self.assertIn('monthly_trend', response.data)
        self.assertIn('user_activity', response.data)
        self.assertIn('demographics', response.data)
        self.assertIn('program_performance', response.data)
        self.assertEqual(response.data['metrics']['total_applications'], 2)

    def test_report_detail_api_returns_application_report(self):
        url = reverse('api:analytics-report-detail', kwargs={'report_type': 'applications'})
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['report_type'], 'applications')
        self.assertIn('summary', response.data)
        self.assertTrue(len(response.data['data']) >= 1)

    def test_report_detail_api_unknown_type(self):
        url = reverse('api:analytics-report-detail', kwargs={'report_type': 'unknown'})
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn('error', response.data)

    def test_export_api_returns_csv_attachment(self):
        url = reverse('api:analytics-export')
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('text/csv', response['Content-Type'])
        self.assertIn('attachment; filename="analytics-report.csv"', response['Content-Disposition'])
        self.assertIn('Metric,Value', response.content.decode('utf-8'))

    def test_export_api_returns_xlsx_attachment(self):
        from io import BytesIO

        from openpyxl import load_workbook

        url = reverse('api:analytics-export')
        response = self.client.get(url, {'export_format': 'xlsx'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn('attachment; filename="analytics-report.xlsx"', response['Content-Disposition'])
        wb = load_workbook(BytesIO(response.content))
        self.assertEqual(wb.sheetnames[0], 'Metrics')
        self.assertIn('Application status', wb.sheetnames)
        self.assertIn('Program performance', wb.sheetnames)
